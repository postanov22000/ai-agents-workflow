import os
import time
import base64
import requests
import io
import json
import smtplib
import ssl
import re
import dns.resolver
import csv
from flask import abort, Flask, render_template, request, redirect, jsonify, make_response, url_for, send_file
from datetime import date, datetime, timezone, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from supabase import create_client, Client
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import Flow
from google.oauth2 import id_token
import google.auth.transport.requests as grequests
from flask_cors import CORS  
from cryptography.fernet import Fernet
from transaction_autopilot import bp as autopilot_bp
from public import public_bp
from io import TextIOWrapper, BytesIO
from openpyxl import load_workbook
from collections import defaultdict
from functools import wraps
from docxtpl import DocxTemplate
import zipfile

# ── single Flask app & blueprint registration ──
app = Flask(__name__, template_folder="templates")
CORS(app, resources={r"/connect-smtp": {"origins": "https://replyzeai.vercel.app"}})
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret")

# ── Env Vars for Admin Catch-All ──
ADMIN_INBOUND_EMAIL = os.environ.get("ADMIN_INBOUND_EMAIL", "replyzeai.inbound@gmail.com")
ADMIN_INBOUND_PASSWORD = os.environ.get("ADMIN_INBOUND_PASSWORD", "") 

# Rate limiting storage - fix structure and initialization
demo_rate_limits = defaultdict(lambda: {
    'emails': {'remaining': 20, 'last_reset': datetime.now()},
    'kits': {'remaining': 20, 'last_reset': datetime.now()},
    'leads': {'remaining': 25, 'last_reset': datetime.now()}
})

# Fixed rate limit decorator
def check_rate_limit(resource):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            ip = request.remote_addr
            now = datetime.now()
            
            # Reset limits based on their specific time periods
            if resource == 'emails':
                # Daily reset for emails
                if (now - demo_rate_limits[ip][resource]['last_reset']).days >= 1:
                    demo_rate_limits[ip][resource]['remaining'] = 20
                    demo_rate_limits[ip][resource]['last_reset'] = now
            else:
                # Monthly reset for kits and leads
                if (now - demo_rate_limits[ip][resource]['last_reset']).days >= 30:
                    if resource == 'kits':
                        demo_rate_limits[ip][resource]['remaining'] = 20
                    else:  # leads
                        demo_rate_limits[ip][resource]['remaining'] = 25
                    demo_rate_limits[ip][resource]['last_reset'] = now
            
            # Check if limit is exceeded
            if demo_rate_limits[ip][resource]['remaining'] <= 0:
                return jsonify({"error": f"{resource.capitalize()} limit exceeded"}), 429
            
            # Decrement the counter and proceed
            demo_rate_limits[ip][resource]['remaining'] -= 1
            return f(*args, **kwargs)
        return decorated_function
    return decorator

#----------------------------------------------------------------------------------
# --- Gmail API Helper Functions ---

def get_gmail_service(user_id):
    """Get Gmail service for a user"""
    try:
        # Get user's Gmail tokens from Supabase
        tok = supabase.table("gmail_tokens") \
                     .select("credentials") \
                     .eq("user_id", user_id) \
                     .single() \
                     .execute()
        
        if not tok.data:
            return None
            
        cd = tok.data[0]["credentials"]
        creds = Credentials(
            token=cd["token"],
            refresh_token=cd["refresh_token"],
            token_uri=cd["token_uri"],
            client_id=cd["client_id"],
            client_secret=cd["client_secret"],
            scopes=cd["scopes"],
        )
        
        # Refresh token if expired
        if creds.expired and creds.refresh_token:
            creds.refresh(GoogleRequest())
            
        # Build Gmail service
        service = build("gmail", "v1", credentials=creds, cache_discovery=False)
        return service
        
    except Exception as e:
        app.logger.error(f"Error getting Gmail service for user {user_id}: {str(e)}")
        return None

def send_email_gmail(user_id, to_email, subject, html_content, cc_emails=None, bcc_emails=None):
    """Send email using Gmail API"""
    try:
        service = get_gmail_service(user_id)
        if not service:
            return False, "Gmail service not available"

        # Create message
        message = MIMEText(html_content, 'html')
        message['to'] = to_email
        message['from'] = "me"  # Gmail API uses 'me' for authenticated user
        message['subject'] = subject
        
        if cc_emails:
            message['cc'] = ', '.join(cc_emails)
        if bcc_emails:
            message['bcc'] = ', '.join(bcc_emails)

        # Encode message
        raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
        
        # Send message
        sent_message = service.users().messages().send(
            userId="me", 
            body={'raw': raw_message}
        ).execute()
        
        app.logger.info(f"Email sent via Gmail API, message ID: {sent_message['id']}")
        return True, "Email sent successfully"
        
    except Exception as e:
        app.logger.error(f"Error sending email via Gmail API: {str(e)}")
        return False, str(e)


def send_email_smtp(from_email, from_password, to_email, subject, body, smtp_host, smtp_port):
    """
    Sends an email using SMTP (SSL).
    """
    msg = MIMEMultipart("alternative")
    msg["From"] = from_email
    msg["To"] = to_email
    msg["Subject"] = subject

    msg.attach(MIMEText(body, "html"))

    # Use SSL for port 465, otherwise use starttls (logic can be expanded if needed)
    if int(smtp_port) == 465:
        with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
            server.login(from_email, from_password)
            server.sendmail(from_email, to_email, msg.as_string())
    else:
        # Fallback for 587 or others
        context = ssl.create_default_context()
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls(context=context)
            server.login(from_email, from_password)
            server.sendmail(from_email, to_email, msg.as_string())


def create_draft_gmail(user_id, to_email, subject, html_content):
    """Create a draft email using Gmail API"""
    try:
        service = get_gmail_service(user_id)
        if not service:
            return False, "Gmail service not available"

        # Create message
        message = MIMEText(html_content, 'html')
        message['to'] = to_email
        message['from'] = "me"
        message['subject'] = subject

        # Encode message
        raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
        
        # Create draft
        draft = service.users().drafts().create(
            userId="me",
            body={'message': {'raw': raw_message}}
        ).execute()
        
        app.logger.info(f"Draft created via Gmail API, draft ID: {draft['id']}")
        return True, "Draft created successfully"
        
    except Exception as e:
        app.logger.error(f"Error creating draft via Gmail API: {str(e)}")
        return False, str(e)

#--------------------------------------------------------------
# Add PlanRateLimiter class definition
class PlanRateLimiter:
    def __init__(self, supabase_client):
        self.supabase = supabase_client
        self.local_cache = defaultdict(dict)

    def _reset_monthly_usage_if_needed(self, user_profile):
        """Reset monthly usage if it's a new month"""
        now = datetime.now(timezone.utc)
        reset_date = user_profile.get('usage_reset_date')
        
        if reset_date:
            if isinstance(reset_date, str):
                reset_date = datetime.fromisoformat(reset_date.replace('Z', '+00:00'))
            
            # Reset on the 1st of each month
            if now.month != reset_date.month or now.year != reset_date.year:
                update_data = {
                    'current_month_leads': 0,
                    'current_month_emails': 0,
                    'current_month_cold_emails': 0,
                    'usage_reset_date': now.isoformat()
                }
                
                self.supabase.table("profiles") \
                    .update(update_data) \
                    .eq("id", user_profile['id']) \
                    .execute()
                
                # Update local profile
                user_profile.update(update_data)
        
        return user_profile
    
    def get_user_plan(self, user_id):
        """Get user's current plan with trial status"""
        try:
            # Check cache first
            if user_id in self.local_cache and 'plan' in self.local_cache[user_id]:
                cached = self.local_cache[user_id]['plan']
                if datetime.now() - cached['fetched_at'] < timedelta(minutes=5):
                    return cached['data']
        
            app.logger.info(f"Getting plan for user {user_id}")
            
            # Get user's profile with plan info
            result = self.supabase.table("profiles") \
                .select("*") \
                .eq("id", user_id) \
                .single() \
                .execute()
            
            if result.data:
                profile = result.data
                
                # Reset monthly usage if needed
                profile = self._reset_monthly_usage_if_needed(profile)
                
                plan_name = profile.get('plan_name', 'starter')
                subscription_status = profile.get('subscription_status', 'active')
                
                # Check if user is in trial period
                trial_ends_at = profile.get('trial_ends_at')
                trial_active = False
                
                if trial_ends_at:
                    trial_ends = datetime.fromisoformat(trial_ends_at.replace('Z', '+00:00'))
                    trial_active = datetime.now(timezone.utc) < trial_ends
                
                if trial_active:
                    # User is in trial
                    plan_data = {
                        'name': plan_name + ' (Trial)',
                        'monthly_leads': profile.get('monthly_leads_limit', 500),
                        'monthly_emails': profile.get('monthly_emails_limit', 500),
                        'connected_accounts': profile.get('connected_accounts_limit', 1),
                        'cold_emails': profile.get('monthly_cold_emails_limit', 200),
                        'document_generation': profile.get('document_generation_enabled', False),
                        'is_trial': True,
                        'trial_days': 14,
                        'trial_ends_at': trial_ends_at,
                        'subscription_status': 'trial',
                        'plan_last_updated': profile.get('plan_last_updated')
                    }
                else:
                    # Regular plan
                    plan_data = {
                        'name': plan_name,
                        'monthly_leads': profile.get('monthly_leads_limit', 500),
                        'monthly_emails': profile.get('monthly_emails_limit', 500),
                        'connected_accounts': profile.get('connected_accounts_limit', 1),
                        'cold_emails': profile.get('monthly_cold_emails_limit', 200),
                        'document_generation': profile.get('document_generation_enabled', False),
                        'is_trial': False,
                        'trial_days': 0,
                        'subscription_status': subscription_status,
                        'plan_last_updated': profile.get('plan_last_updated')
                    }
                
                # Add current usage from profile
                plan_data.update({
                    'current_leads': profile.get('current_month_leads', 0),
                    'current_emails': profile.get('current_month_emails', 0),
                    'current_cold_emails': profile.get('current_month_cold_emails', 0)
                })
                
                # Cache the result
                self.local_cache[user_id]['plan'] = {
                    'data': plan_data,
                    'fetched_at': datetime.now()
                }
                
                return plan_data
            
            # No profile found - default
            default_plan = {
                'name': 'Starter',
                'monthly_leads': 500,
                'monthly_emails': 500,
                'connected_accounts': 1,
                'cold_emails': 200,
                'document_generation': False,
                'is_trial': False,
                'trial_days': 0,
                'subscription_status': 'active',
                'current_leads': 0,
                'current_emails': 0,
                'current_cold_emails': 0
            }
            
            return default_plan
            
        except Exception as e:
            app.logger.error(f"Error getting user plan: {str(e)}")
            fallback = {
                'name': 'Starter',
                'monthly_leads': 500,
                'monthly_emails': 500,
                'connected_accounts': 1,
                'cold_emails': 200,
                'document_generation': False,
                'is_trial': False,
                'trial_days': 0,
                'subscription_status': 'active',
                'current_leads': 0,
                'current_emails': 0,
                'current_cold_emails': 0
            }
            return fallback
    
    def check_rate_limit(self, user_id, resource_type, amount=1):
        """Check if user has exceeded rate limit for a resource"""
        try:
            # First, get the latest plan data
            plan = self.get_user_plan(user_id)
            
            # Map resource types to plan limits
            resource_map = {
                'leads': ('monthly_leads', 'current_leads'),
                'emails': ('monthly_emails', 'current_emails'),
                'cold_emails': ('cold_emails', 'current_cold_emails'),
                'connected_accounts': ('connected_accounts', None)
            }
            
            if resource_type not in resource_map:
                return False, 0, f"Unknown resource type: {resource_type}"
            
            limit_key, current_key = resource_map[resource_type]
            plan_limit = plan.get(limit_key, 0)
            current_usage = plan.get(current_key, 0) if current_key else 0
            
            app.logger.info(f"Rate limit check for user {user_id}: {resource_type}")
            app.logger.info(f"  Plan limit: {plan_limit}, Current usage: {current_usage}, Requested: {amount}")
            
            # Check if adding amount would exceed limit
            if current_usage + amount > plan_limit:
                remaining = max(0, plan_limit - current_usage)
                message = f"{resource_type.replace('_', ' ').title()} limit exceeded. Plan limit: {plan_limit}, Used: {current_usage}"
                app.logger.warning(f"Rate limit exceeded: {message}")
                return False, remaining, message
            
            # If allowed, return success
            remaining = plan_limit - current_usage
            return True, remaining, f"Limit: {plan_limit}, Used: {current_usage}, Remaining: {remaining}"
            
        except Exception as e:
            app.logger.error(f"Error checking rate limit: {str(e)}", exc_info=True)
            return False, 0, f"Error checking limits: {str(e)}"
    
    def _increment_usage(self, user_id, resource_type, amount=1):
        """Increment usage counter in database"""
        try:
            # Map resource types to column names
            column_map = {
                'leads': 'current_month_leads',
                'emails': 'current_month_emails',
                'cold_emails': 'current_month_cold_emails'
            }
        
            if resource_type not in column_map:
                return
        
            column = column_map[resource_type]
            
            # Use RPC function to increment or direct update
            try:
                self.supabase.rpc('increment_usage', {
                    'user_id': user_id,
                    'column_name': column,
                    'amount': amount
                }).execute()
            except Exception as rpc_error:
                # Fallback to direct update
                app.logger.warning(f"RPC increment failed, using direct update: {str(rpc_error)}")
                # Get current value first
                result = self.supabase.table("profiles") \
                    .select(column) \
                    .eq("id", user_id) \
                    .single() \
                    .execute()
                
                current_value = result.data.get(column, 0) if result.data else 0
                new_value = current_value + amount
            
                # Update the value
                self.supabase.table("profiles") \
                    .update({column: new_value}) \
                    .eq("id", user_id) \
                    .execute()
            
        except Exception as e:
            app.logger.error(f"Error incrementing usage: {str(e)}")
#----------------------------------------------------------------    
@app.route("/signin2")
def signin():
    user_id = request.args.get("user_id", "")
    return render_template("signin2.html", user_id=user_id)
#--------------------------------------------------------------
app.register_blueprint(autopilot_bp, url_prefix="/autopilot")
app.register_blueprint(public_bp)

# --- Supabase setup ---
SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_ANON_KEY = os.environ["SUPABASE_ANON_KEY"]
SUPABASE_SERVICE_ROLE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
supabase: Client = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
SUPABASE_SERVICE: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
# Edge Function base URL *without* trailing slash or endpoint
EDGE_BASE_URL = os.environ.get("EDGE_BASE_URL", "").rstrip("/")
ENCRYPTION_KEY = os.environ["ENCRYPTION_KEY"].encode()  # 32-url-safe-base64 bytes
fernet = Fernet(ENCRYPTION_KEY)
# Retry configuration for calling the Edge Function
MAX_RETRIES = 5
RETRY_BACKOFF_BASE = 2

# Define follow-up sequence (days after initial contact)

rate_limiter = PlanRateLimiter(supabase)

FOLLOW_UP_SEQUENCE = [
    {"delay_days": 0, "name": "Immediate Follow-up"},
    {"delay_days": 1, "name": "Day 1 Follow-up"},
    {"delay_days": 3, "name": "Day 3 Follow-up"},
    {"delay_days": 7, "name": "Day 7 Follow-up"},
    {"delay_days": 14, "name": "Day 14 Follow-up"},
    {"delay_days": 30, "name": "Day 30 Follow-up"},
]

#----------------------------------------------------------------------------
# --- Helper Functions for Centralized Forwarding ---
def normalize_display_name(display_name):
    """Normalize display name to create a clean username for email alias"""
    if not display_name:
        return "user"
    # Remove all non-alphanumeric characters and convert to lowercase (or keep case if preferred)
    # Keeping case can be nice for readability (JohnDoe) but lowercase is safer for email.
    # We will just strip spaces and special chars.
    clean_name = re.sub(r'[^a-zA-Z0-9]', '', display_name)
    return clean_name

# ---------------------------------------------------------------------------
def call_edge(endpoint_path: str, payload: dict, return_response: bool = False):
    url = f"{EDGE_BASE_URL}{endpoint_path}"
    app.logger.info(f"🔗 call_edge → URL: {url}")
    app.logger.info(f"🔗 call_edge → Payload: {payload}")

    headers = {
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "apikey":        SUPABASE_SERVICE_ROLE_KEY,
        "Content-Type":  "application/json"
    }

    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.post(url, json=payload, headers=headers, timeout=120)
            app.logger.info(f"↩️  Response [{resp.status_code}]: {resp.text}")

            if resp.status_code == 200:
                if return_response:
                    return resp
                else:
                    return True
            elif resp.status_code == 429:
                wait = RETRY_BACKOFF_BASE ** attempt
                app.logger.warning(f"[{endpoint_path}] Rate‐limited, retry {attempt+1}/{MAX_RETRIES} after {wait}s")
                time.sleep(wait)
                continue
            else:
                app.logger.error(f"[{endpoint_path}] Failed ({resp.status_code}): {resp.text}")
                if return_response:
                    return resp
                else:
                    return False
        except requests.RequestException as e:
            wait = RETRY_BACKOFF_BASE ** attempt
            app.logger.error(f"[{endpoint_path}] Exception: {e}, retrying in {wait}s")
            time.sleep(wait)
    app.logger.error(f"[{endpoint_path}] Exceeded max retries.")
    if return_response:
        return None
    else:
        return False

# ── Routes ──
#-----------------------------------------------


# Add this near the top of your app.py after creating the Flask app
@app.template_filter('format_date')
def format_date_filter(value):
    if not value:
        return ""
    try:
        # Try to parse the date string
        date_obj = datetime.fromisoformat(value.replace('Z', '+00:00'))
        return date_obj.strftime("%b %d, %Y %I:%M %p")
    except:
        return value





#-------------------------------------------------
from flask import url_for

@app.route("/")
def home():
    """
    Just redirect to /dashboard, passing along user_id if any.
    """
    user_id = request.args.get("user_id", "")
    # Redirect to /dashboard?user_id=<...> (blank if none)
    return redirect(f"/dashboard?user_id={user_id}")



@app.route("/dashboard")
def dashboard():
    user_id = request.args.get("user_id", "").strip()

    # ── GUEST DEFAULTS ──
    name = "Guest"
    ai_enabled = False
    generate_leases = False
    emails_sent = 0
    time_saved = 0
    show_reconnect = False
    revenue = 0
    revenue_change = 0
    email_mode = "auto"

    # Ensure these always exist for the template
    kits_generated = 0
    estimated_saved = 0
    needs_mode_selection = False

    if user_id:
        # 1) Load profile and check email mode
        try:
            profile_resp = (
                supabase.table("profiles")
                .select("full_name, ai_enabled, generate_leases, email_mode")
                .eq("id", user_id)
                .single()
                .execute()
            )
            if profile_resp.data:
                profile_data = profile_resp.data
                name = profile_data["full_name"]
                ai_enabled = profile_data["ai_enabled"]
                generate_leases = profile_data["generate_leases"]
                email_mode = profile_data.get("email_mode", "auto")
                
                # Check if user needs to select email mode
                if profile_data.get("email_mode") is None:
                    needs_mode_selection = True
                    
        except Exception as e:
            app.logger.warning(f"dashboard: failed to load profile for {user_id}: {str(e)}")
            needs_mode_selection = True

        # 2) FIXED: Count emails sent BY this user (using sent_by_account)
        try:
            today = date.today().isoformat()
            rows = (
                supabase.table("emails")
                .select("sent_at, sent_by_account, original_user_id")
                .eq("sent_by_account", user_id)  # Count emails sent BY this user's account
                .eq("status", "sent")
                .execute()
                .data or []
            )
            # Count all emails sent by this user's account
            emails_sent = len(rows)
            time_saved = emails_sent * 5.5
        except Exception:
            app.logger.warning(f"dashboard: failed to count emails for {user_id}")
            emails_sent = 0
            time_saved = 0

        # 3) Gmail reconnect flag - only show for auto mode
        show_reconnect = False
        if email_mode == "auto":
            try:
                toks = (
                    supabase.table("gmail_tokens")
                    .select("credentials")
                    .eq("user_id", user_id)
                    .execute()
                    .data or []
                )
                if toks:
                    cd = toks[0]["credentials"]
                    creds = Credentials(
                        token=cd["token"],
                        refresh_token=cd["refresh_token"],
                        token_uri=cd["token_uri"],
                        client_id=cd["client_id"],
                        client_secret=cd["client_secret"],
                        scopes=cd["scopes"],
                    )
                    show_reconnect = creds.expired
            except Exception:
                app.logger.warning(f"dashboard: failed to check Gmail token for {user_id}")

        # 4) Count "kits generated" for this user
        kit_rows = (
            supabase.table("transactions")
            .select("id")
            .eq("user_id", user_id)
            .eq("kit_generated", True)
            .execute()
            .data or []
        )
        kits_generated = len(kit_rows)

        # 5) Compute extra estimated time saved (e.g. 15 min per kit)
        PER_KIT_SAVE_MINUTES = 15
        estimated_saved = kits_generated * PER_KIT_SAVE_MINUTES

    # If HTMX request and needs mode selection, return just the modal
    if request.headers.get('HX-Request') and needs_mode_selection:
        return render_template("mode_selection_modal.html", user_id=user_id)

    # ── Render dashboard ──
    return render_template(
        "dashboard.html",
        user_id=user_id,
        name=name,
        ai_enabled=ai_enabled,
        generate_leases=generate_leases,
        emails_sent=emails_sent,
        time_saved=time_saved,
        estimated_saved=estimated_saved,
        kits_generated=kits_generated,
        show_reconnect=show_reconnect,
        revenue=revenue,
        revenue_change=revenue_change,
        email_mode=email_mode,
        needs_mode_selection=needs_mode_selection
    )
#--------------------------------------------------------------------------------------------------------------
@app.route("/dashboard/leads")
def dashboard_leads():
    user_id = _require_user()
    return render_template("partials/leads_funnel.html", user_id=user_id)

# Fix for the search error - update the leads_list function
@app.route("/dashboard/leads/list")
def leads_list():
    user_id = _require_user()
    filter_type = request.args.get("filter", "all")
    search_query = request.args.get("q", "")
    
    # Build query based on filters
    query = supabase.table("leads").select("*").eq("user_id", user_id)
    
    if filter_type != "all":
        query = query.eq("status", filter_type)
    
    # Execute query first to get all results
    try:
        result = query.execute()
        leads = result.data or []
    except Exception as e:
        app.logger.error(f"Error fetching leads: {str(e)}")
        leads = []
    
    # Apply search filter in Python
    if search_query:
        search_lower = search_query.lower()
        leads = [lead for lead in leads if 
                (lead.get("first_name", "").lower().find(search_lower) != -1 or
                 lead.get("last_name", "").lower().find(search_lower) != -1 or
                 lead.get("email", "").lower().find(search_lower) != -1 or
                 lead.get("brokerage", "").lower().find(search_lower) != -1)]
    
    # Calculate funnel counts
    counts = {
        "new": 0,
        "contacted": 0,
        "proposal": 0,
        "closed": 0
    }
    
    try:
        # Get counts for each status
        for status in counts.keys():
            count_result = supabase.table("leads").select("id", count="exact").eq("user_id", user_id).eq("status", status).execute()
            counts[status] = count_result.count or 0
    except Exception as e:
        app.logger.error(f"Error counting leads by status: {str(e)}")
    
    return render_template("partials/leads_list.html", leads=leads, counts=counts, user_id=user_id)


@app.route("/dashboard/leads/search")
def search_leads():
    # Reuse the leads_list function but with search parameters
    return leads_list()

@app.route("/dashboard/leads/<lead_id>/view")
def view_lead(lead_id):
    user_id = _require_user()
    
    try:
        # Get lead details
        lead = supabase.table("leads").select("*").eq("id", lead_id).eq("user_id", user_id).single().execute().data
        
        # Get follow-up history
        follow_ups = supabase.table("lead_follow_ups").select("*").eq("lead_id", lead_id).order("scheduled_at").execute().data or []
        
        return render_template("partials/lead_detail.html", lead=lead, follow_ups=follow_ups, user_id=user_id)
    except Exception as e:
        app.logger.error(f"Error fetching lead details: {str(e)}")
        return "<div class='error'>Error loading lead details: Missing required database columns</div>", 500

@app.route("/dashboard/leads/<lead_id>/update-status", methods=["POST"])
def update_lead_status(lead_id):
    user_id = _require_user()
    new_status = request.form.get("status")
    
    if not new_status:
        return jsonify({"error": "Status is required"}), 400
    
    try:
        # Update lead status
        supabase.table("leads").update({
            "status": new_status,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }).eq("id", lead_id).eq("user_id", user_id).execute()
        
        return "", 204
    except Exception as e:
        app.logger.error(f"Error updating lead status: {str(e)}")
        return jsonify({"error": "Failed to update status"}), 500

# Fix for the lead notes error - update the add_lead_note function
@app.route("/dashboard/leads/<lead_id>/add-note", methods=["POST"])
def add_lead_note(lead_id):
    user_id = _require_user()
    note_content = request.form.get("note")
    
    if not note_content:
        return jsonify({"error": "Note content is required"}), 400
    
    try:
        # First verify the lead exists and belongs to this user
        lead_check = supabase.table("leads").select("id").eq("id", lead_id).eq("user_id", user_id).execute()
        if not lead_check.data:
            return jsonify({"error": "Lead not found or access denied"}), 404
        
        # Add note to lead
        result = supabase.table("lead_notes").insert({
            "lead_id": lead_id,
            "user_id": user_id,
            "content": note_content,
            "created_at": datetime.now(timezone.utc).isoformat()
        }).execute()
        
        # Check if insertion was successful
        if not result.data:
            app.logger.error(f"Note insertion failed: {result}")
            return jsonify({"error": "Failed to add note - no data returned"}), 500
            
        return "", 204
    except Exception as e:
        app.logger.error(f"Error adding lead note: {str(e)}", exc_info=True)
        
        # Check if it's a specific API error
        error_msg = str(e)
        if "foreign key constraint" in error_msg.lower():
            return jsonify({"error": "Invalid lead ID"}), 400
        elif "null value" in error_msg.lower():
            return jsonify({"error": "Missing required fields"}), 400
            
        return jsonify({"error": "Failed to add note"}), 500


@app.route("/dashboard/leads/export")
def export_leads():
    user_id = _require_user()
    filter_type = request.args.get("filter", "all")
    
    try:
        # Build query
        query = supabase.table("leads").select("*").eq("user_id", user_id)
        
        if filter_type != "all":
            query = query.eq("status", filter_type)
        
        leads = query.execute().data or []
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(["First Name", "Last Name", "Email", "Brokerage", "Service", "City", "Status", "Last Contact"])
        
        # Write data
        for lead in leads:
            writer.writerow([
                lead.get("first_name", ""),
                lead.get("last_name", ""),
                lead.get("email", ""),
                lead.get("brokerage", ""),
                lead.get("service", ""),
                lead.get("city", ""),
                lead.get("status", "new"),
                lead.get("last_contacted_at", "")
            ])
        
        # Prepare response
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response.headers["Content-type"] = "text/csv"
        
        return response
    except Exception as e:
        app.logger.error(f"Error exporting leads: {str(e)}")
        return jsonify({"error": "Failed to export leads"}), 500
#------------------------------------------------------------------------------------------------------------
@app.route("/dashboard/new_transaction")
def dashboard_new_transaction():
    user_id = request.args.get("user_id") or abort(401)
    return render_template("partials/new_transaction.html", user_id=user_id)
  
@app.route("/dashboard/responded_emails")
def dashboard_responded_emails():
    user_id = request.args.get("user_id") or abort(401)
    # Select emails for this user that were sent/drafted and that have an original_content field
    try:
        emails = (
            supabase.table("emails")
                    .select("id, sender_email, subject, original_content, status, sent_at")
                    .eq("user_id", user_id)
                    .in_("status", ["sent","drafted"])   # treat drafted as 'responded' if you want
                    .order("sent_at", desc=True)
                    .execute()
                    .data
            or []
        )
    except Exception:
        app.logger.exception("failed to load responded emails")
        emails = []

    return render_template("partials/responded_emails.html", emails=emails, user_id=user_id)


@app.route("/dashboard/email/<email_id>")
def dashboard_email_view(email_id):
    """Return a small partial showing full original_content — HTMX call for modal."""
    try:
        rec = supabase.table("emails").select("*").eq("id", email_id).single().execute().data
    except Exception:
        rec = None

    if not rec:
        return "<div class='chart-container'>Email not found.</div>", 404

    return render_template("partials/email_modal.html", email=rec)


@app.route("/dashboard/analytics")
def dashboard_analytics():
    user_id = _require_user()
    return render_template("partials/analytics.html", user_id=user_id)

@app.route("/dashboard/users")
def dashboard_users():
    user_id = _require_user()
    users = supabase.table("profiles").select("id, full_name, email").execute().data or []
    return render_template("partials/users.html", users=users)

@app.route("/dashboard/billing")
def dashboard_billing():
    user_id = _require_user()
    return render_template("partials/billing.html", user_id=user_id)

# Update the settings route to remove SMTP references
@app.route("/dashboard/settings", methods=["GET", "POST"])
def dashboard_settings():
    user_id = _require_user()

    if request.method == "POST":
        section = request.form.get("section")
        if section == "profile":
            new_display_name = request.form.get("display_name", "").strip()
            new_signature = request.form.get("signature", "").strip()
            supabase.table("profiles").update({
                "display_name": new_display_name,
                "signature": new_signature
            }).eq("id", user_id).execute()

    # Fetch profile & flags
    profile_resp = supabase.table("profiles") \
                           .select("display_name, signature, ai_enabled") \
                           .eq("id", user_id) \
                           .single() \
                           .execute()
    
    profile = profile_resp.data or {
        "display_name": "",
        "signature": "",
        "ai_enabled": False
    }

    # Check Gmail connection status
    gmail_connected = False
    show_reconnect = False
    
    try:
        service = get_gmail_service(user_id)
        if service:
            gmail_connected = True
            # Check if token needs refresh by trying a simple operation
            # If it fails, show reconnect button
            try:
                # Simple check to see if service works
                service.users().getProfile(userId='me').execute()
            except Exception:
                show_reconnect = True
    except Exception:
        app.logger.warning(f"settings: could not check Gmail token for {user_id}")

    # Render template
    return render_template(
        "partials/settings.html",
        profile=profile,
        user_id=user_id,
        gmail_connected=gmail_connected,
        show_reconnect=show_reconnect
    )



@app.route("/dashboard/home")
def dashboard_home():
    user_id = request.args.get("user_id")
    if not user_id:
        return "Missing user_id", 401

    # (Same logic as /dashboard for HTMX partial)
    profile_resp = (
        supabase.table("profiles")
                .select("display_name, ai_enabled, email, generate_leases, email_mode")
                .eq("id", user_id)
                .single()
                .execute()
    )
    if profile_resp.data is None:
        return "Profile query error", 500

    profile         = profile_resp.data
    full_name       = profile.get("display_name", "")
    ai_enabled      = profile.get("ai_enabled", True)
    generate_leases = profile.get("generate_leases", False)
    email_mode = profile.get("email_mode")  # Get email_mode

    # FIXED: Count emails sent BY this user (using sent_by_account)
    today     = date.today().isoformat()
    sent_rows = (
        supabase.table("emails")
                .select("sent_at, sent_by_account, original_user_id")
                .eq("sent_by_account", user_id)  # Emails sent BY this user's account
                .eq("status", "sent")
                .execute()
                .data
        or []
    )
    # Count total emails sent by this user's account (all time)
    emails_sent_total = len(sent_rows)
    
    # Count today's emails specifically
    emails_sent_today = sum(1 for e in sent_rows if e.get("sent_at", "").startswith(today))
    
    # Use total for time saved calculation
    time_saved = emails_sent_total * 5.5

    token_rows = (
        supabase.table("gmail_tokens")
                .select("credentials")
                .eq("user_id", user_id)
                .execute()
                .data
        or []
    )
    show_reconnect = True
    if token_rows:
        creds_data = token_rows[0]["credentials"]
        try:
            creds = Credentials(
                token=creds_data["token"],
                refresh_token=creds_data["refresh_token"],
                token_uri=creds_data["token_uri"],
                client_id=creds_data["client_id"],
                client_secret=creds_data["client_secret"],
                scopes=cd["scopes"],
            )
            show_reconnect = creds.expired
        except Exception:
            pass
            
    # 4) Count "kits generated" for this user
    kit_rows = (
        supabase
        .table("transactions")
        .select("id")
        .eq("user_id", user_id)
        .eq("kit_generated", True)
        .execute()
        .data
        or []
    )
    kits_generated = len(kit_rows)

    # 5) Compute extra estimated time saved
    PER_KIT_SAVE_MINUTES = 15
    estimated_saved = kits_generated * PER_KIT_SAVE_MINUTES

    return render_template(
        "partials/home.html",
        name=full_name,
        user_id=user_id,
        emails_sent=emails_sent_total,  # Use total count for display
        time_saved=time_saved,
        estimated_saved=estimated_saved,
        kits_generated=kits_generated,
        ai_enabled=ai_enabled,
        show_reconnect=show_reconnect,
        generate_leases=generate_leases,
        email_mode=email_mode
    )
#----------------------------------------------------------------------
@app.route("/reconnect_gmail")
def reconnect_gmail():
    """Handles both initial connection and reconnection to Gmail"""
    user_id = request.args.get("user_id")
    if not user_id:
        return "Missing user ID", 400

    # Validate user_id format before proceeding
    if not is_valid_uuid(user_id):
        app.logger.warning(f"Invalid user_id format: {user_id}")
        return "Invalid user ID format", 400

    flow = Flow.from_client_config(
        {
            "web": {
                "client_id": os.environ["GOOGLE_CLIENT_ID"],
 
                "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [os.environ["REDIRECT_URI"]]
            }
        },
        scopes=[
            "https://www.googleapis.com/auth/gmail.send",
            "https://www.googleapis.com/auth/userinfo.email",
            "openid"
        ]
    )
    flow.redirect_uri = os.environ["REDIRECT_URI"]
    authorization_url, _ = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
        state=user_id
    )
    return redirect(authorization_url)
  

#------------------------------------------ 


#------------------------------------------
# Update your send_email function to require valid connection
#@app.route("/send", methods=["POST"])
#def send_email():
#    data = request.get_json()
#    user_id = data["user_id"]
#    
#    # Check email connection before proceeding
#    require_valid_email_connection(user_id)
#    
#    # Rest of your send logic...
#    to = data["to"]
#    subject = data["subject"]
#    body = data["body"]
#
#    # else: your existing Gmail-API‐based fetch
#    return fetch_via_gmail_api(user_id)

# Add this route to update the connection status in the database
#@app.route("/update_connection_status", methods=["POST"])
#def update_connection_status():
#    user_id = request.form.get("user_id")
#    status = request.form.get("status")  # "valid" or "invalid"
#    
#    if not user_id or not status:
#        return jsonify({"status": "error", "message": "Missing parameters"}), 400
#    
#    # Update the connection status in the database
#    supabase.table("profiles").update({
#        "email_connection_status": status,
#        "connection_checked_at": datetime.now(timezone.utc).isoformat()
#    }).eq("id", user_id).execute()
#    
#    return jsonify({"status": "success"})
#-----------------------------------------------------------------------------
#polling gmail
@app.route("/connect_gmail2")
def connect_gmail2():
    """
    Initiates Gmail OAuth flow.
    """
    user_id = request.args.get("user_id")
    if not user_id or not is_valid_uuid(user_id):
        app.logger.error(f"Missing or invalid user_id in connect_gmail: {user_id}")
        return "Missing or invalid user ID", 400

    flow = Flow.from_client_config(
        {
            "web": {
                "client_id": os.environ["GOOGLE_CLIENT_ID2"],
                "client_secret": os.environ["GOOGLE_CLIENT_SECRET2"],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [os.environ["REDIRECT_URI2"]]
            }
        },
        scopes=[
            "https://www.googleapis.com/auth/gmail.readonly",
            "https://www.googleapis.com/auth/userinfo.email",
            "openid"
        ]
    )
    flow.redirect_uri = os.environ["REDIRECT_URI2"]

    # ✅ now user_id is safely defined
    authorization_url, _ = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
        state=user_id
    )

    return redirect(authorization_url)



import uuid

def is_valid_uuid(val: str) -> bool:
    try:
        uuid.UUID(str(val))
        return True
    except ValueError:
        return False


@app.route("/oauth2callback2")
def oauth2callback2():
    """Handles OAuth2 callback from Google"""
    try:
        # Extract state parameter containing user_id
        user_id = request.args.get("state")
        if not user_id:
            app.logger.error("OAuth2 callback missing state parameter")
            return "<h1>Authentication Failed</h1><p>Missing state parameter</p>", 400

        # ✅ Validate UUID format before querying Supabase
        if not is_valid_uuid(user_id):
            app.logger.error(f"Invalid user_id format in state: {user_id}")
            return "<h1>Authentication Failed</h1><p>Invalid user ID format</p>", 400

        # Check if user exists in Supabase
        try:
            user_check = supabase.table("profiles").select("id").eq("id", user_id).execute()
            if not user_check.data:
                app.logger.error(f"User not found: {user_id}")
                return "<h1>Authentication Failed</h1><p>User not found</p>", 400
        except Exception as e:
            app.logger.error(f"Error checking user: {str(e)}")
            return "<h1>Authentication Failed</h1><p>Error validating user</p>", 500

        # Continue OAuth flow
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": os.environ["GOOGLE_CLIENT_ID2"],
                    "client_secret": os.environ["GOOGLE_CLIENT_SECRET2"],
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [os.environ["REDIRECT_URI2"]]
                }
            },
            scopes=[
                "https://www.googleapis.com/auth/gmail.readonly",
                "https://www.googleapis.com/auth/userinfo.email",
                "openid"
            ],
            state=user_id
        )
        flow.redirect_uri = os.environ["REDIRECT_URI2"]
        flow.fetch_token(authorization_response=request.url)
        credentials = flow.credentials

        # Verify ID token
        id_info = id_token.verify_oauth2_token(
            credentials.id_token,
            grequests.Request(),
            os.environ["GOOGLE_CLIENT_ID2"]
        )
        email = id_info.get("email")
        if not email:
            raise ValueError("No email found in Google ID token")

        # Upsert Gmail tokens
        creds_payload = {
            "user_id": user_id,
            "user_email": email,
            "credentials": {
                "token": credentials.token,
                "refresh_token": credentials.refresh_token,
                "token_uri": credentials.token_uri,
                "client_id": credentials.client_id,
                "client_secret": credentials.client_secret,
                "scopes": credentials.scopes
            }
        }

        try:
            supabase.table("gmail_tokens2").upsert(creds_payload).execute()
        except Exception as db_error:
            app.logger.error(f"Database error during token upsert: {str(db_error)}")
            if "uuid" in str(db_error).lower() and "format" in str(db_error).lower():
                app.logger.warning(f"Non-UUID user_id detected: {user_id}")
                return "<h1>Authentication Failed</h1><p>User ID format issue. Please contact support.</p>", 400
            else:
                raise db_error

        # Update user profile
        full_name = id_info.get("name") or email.split("@")[0]
        supabase.table("profiles").update({
            "email": email,
            "full_name": full_name,
            "ai_enabled": True
        }).eq("id", user_id).execute()

        return redirect(f"/dashboard?user_id={user_id}")

    except Exception as e:
        app.logger.error(f"OAuth2 Callback Error: {str(e)}", exc_info=True)
        return f"<h1>Authentication Failed</h1><p>{str(e)}</p>", 500
#-----------------------------------------------------------------------
@app.route("/connect_gmail")
def connect_gmail():
    """
    Initiates Gmail OAuth flow.
    """
    user_id = request.args.get("user_id")
    if not user_id or not is_valid_uuid(user_id):
        app.logger.error(f"Missing or invalid user_id in connect_gmail: {user_id}")
        return "Missing or invalid user ID", 400

    flow = Flow.from_client_config(
        {
            "web": {
                "client_id": os.environ["GOOGLE_CLIENT_ID"],
                "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [os.environ["REDIRECT_URI"]]
            }
        },
        scopes=[
            "https://www.googleapis.com/auth/gmail.send",
            "https://www.googleapis.com/auth/userinfo.email",
            "openid"
        ]
    )
    flow.redirect_uri = os.environ["REDIRECT_URI"]

    # ✅ now user_id is safely defined
    authorization_url, _ = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
        state=user_id
    )

    return redirect(authorization_url)



import uuid

def is_valid_uuid(val: str) -> bool:
    try:
        uuid.UUID(str(val))
        return True
    except ValueError:
        return False


@app.route("/oauth2callback")
def oauth2callback():
    """Handles OAuth2 callback from Google"""
    try:
        # Extract state parameter containing user_id
        user_id = request.args.get("state")
        if not user_id:
            app.logger.error("OAuth2 callback missing state parameter")
            return "<h1>Authentication Failed</h1><p>Missing state parameter</p>", 400

        # ✅ Validate UUID format before querying Supabase
        if not is_valid_uuid(user_id):
            app.logger.error(f"Invalid user_id format in state: {user_id}")
            return "<h1>Authentication Failed</h1><p>Invalid user ID format</p>", 400

        # Check if user exists in Supabase
        try:
            user_check = supabase.table("profiles").select("id").eq("id", user_id).execute()
            if not user_check.data:
                app.logger.error(f"User not found: {user_id}")
                return "<h1>Authentication Failed</h1><p>User not found</p>", 400
        except Exception as e:
            app.logger.error(f"Error checking user: {str(e)}")
            return "<h1>Authentication Failed</h1><p>Error validating user</p>", 500

        # Continue OAuth flow
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": os.environ["GOOGLE_CLIENT_ID"],
                    "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [os.environ["REDIRECT_URI"]]
                }
            },
            scopes=[
                "https://www.googleapis.com/auth/gmail.send",
                "https://www.googleapis.com/auth/userinfo.email",
                "openid"
            ],
            state=user_id
        )
        flow.redirect_uri = os.environ["REDIRECT_URI"]
        flow.fetch_token(authorization_response=request.url)
        credentials = flow.credentials

        # Verify ID token
        id_info = id_token.verify_oauth2_token(
            credentials.id_token,
            grequests.Request(),
            os.environ["GOOGLE_CLIENT_ID"]
        )
        email = id_info.get("email")
        if not email:
            raise ValueError("No email found in Google ID token")

        # Upsert Gmail tokens
        creds_payload = {
            "user_id": user_id,
            "user_email": email,
            "credentials": {
                "token": credentials.token,
                "refresh_token": credentials.refresh_token,
                "token_uri": credentials.token_uri,
                "client_id": credentials.client_id,
                "client_secret": credentials.client_secret,
                "scopes": credentials.scopes
            }
        }

        try:
            supabase.table("gmail_tokens").upsert(creds_payload).execute()
        except Exception as db_error:
            app.logger.error(f"Database error during token upsert: {str(db_error)}")
            if "uuid" in str(db_error).lower() and "format" in str(db_error).lower():
                app.logger.warning(f"Non-UUID user_id detected: {user_id}")
                return "<h1>Authentication Failed</h1><p>User ID format issue. Please contact support.</p>", 400
            else:
                raise db_error

        # Update user profile
        full_name = id_info.get("name") or email.split("@")[0]
        supabase.table("profiles").update({
            "email": email,
            "full_name": full_name,
            "ai_enabled": True
        }).eq("id", user_id).execute()

        return redirect(f"/dashboard?user_id={user_id}")

    except Exception as e:
        app.logger.error(f"OAuth2 Callback Error: {str(e)}", exc_info=True)
        return f"<h1>Authentication Failed</h1><p>{str(e)}</p>", 500

      
@app.route("/complete_profile", methods=["GET", "POST"])
def complete_profile():
    user_id = request.args.get("user_id")
    if not user_id:
        return "Missing user_id", 401

    if request.method == "POST":
        display_name = request.form.get("display_name", "").strip()
        signature    = request.form.get("signature", "").strip()

        supabase.table("profiles") \
            .update({"display_name": display_name,
                     "signature": signature}) \
            .eq("id", user_id) \
            .execute()

        return redirect(f"/dashboard?user_id={user_id}")

    return render_template("complete_profile.html", user_id=user_id)

@app.route("/disconnect_gmail", methods=["POST"])
def disconnect_gmail():
    user_id = request.form.get("user_id")
    supabase.table("gmail_tokens").delete().eq("user_id", user_id).execute()
    return redirect(f"/dashboard?user_id={user_id}")

def _require_user():
    uid = request.args.get("user_id") or request.form.get("user_id")
    if not uid:
        abort(401, "Missing user_id")
    return uid

@app.route("/new_lease", methods=["GET"])
def new_lease_form():
    user_id = _require_user()
    return render_template("new_lease.html", user_id=user_id)

@app.route("/new_lease", methods=["POST"])
def new_lease_submit():
    user_id = _require_user()

    data = {
        "property_name":    request.form["propertyName"],
        "property_type":    request.form["propertyType"],
        "address":          request.form["address"],
        "suite":            request.form.get("suite",""),
        "square_feet":      request.form["squareFeet"],
        "tenant_name":      request.form["tenantName"],
        "tenant_type":      request.form["tenantType"],
        "lease_type":       request.form["leaseType"],
        "lease_term":       request.form["leaseTerm"],
        "start_date":       request.form["startDate"],
        "end_date":         request.form["endDate"],
        "base_rent":        request.form["baseRent"],
        "annual_increase":  request.form.get("annualIncrease",""),
        "security_deposit": request.form.get("securityDeposit",""),
        "parking_spaces":   request.form.get("parkingSpaces",""),
        "parking_fee":      request.form.get("parkingFee",""),
        "additional_terms": request.form.get("additionalTerms",""),
        "tenant_improvements": "Yes" if request.form.get("tenantImprovements") else "No",
        "renewal_option":      "Yes" if request.form.get("renewalOption") else "No",
        "exclusive_use":       "Yes" if request.form.get("exclusiveUse") else "No",
    }

    html_body = f"""
    <html><body>
      <h2>Lease Agreement</h2>
      <p><strong>Property:</strong> {data['property_name']} ({data['property_type'].title()})<br>
      <strong>Address:</strong> {data['address']} Suite {data['suite']}<br>
      <strong>Size:</strong> {data['square_feet']} sqft</p>

      <h3>Tenant</h3>
      <p>{data['tenant_name']} ({data['tenant_type'].title()})</p>

      <h3>Terms</h3>
      <p><strong>Type:</strong> {data['lease_type'].replace('-', '').title()}<br>
      <strong>Term:</strong> {data['lease_term']} months<br>
      <strong>Dates:</strong> {data['start_date']} → {data['end_date']}</p>

      <h3>Financials</h3>
      <p><strong>Base Rent:</strong> ${data['base_rent']} per sqft/yr<br>
      <strong>Annual Increase:</strong> {data['annual_increase']}%<br>
      <strong>Security Deposit:</strong> ${data['security_deposit']}<br>
      <strong>Parking:</strong> {data['parking_spaces']} spaces @ ${data['parking_fee']}/mo</p>

      <h3>Additional Terms</h3>
      <p>{data['additional_terms']}</p>
      <ul>
        <li>Tenant Improvements: {data['tenant_improvements']}</li>
        <li>Renewal Option: {data['renewal_option']}</li>
        <li>Exclusive Use Clause: {data['exclusive_use']}</li>
      </ul>
    </body></html>
    """

    tok = (supabase.table("gmail_tokens")
                .select("credentials")
                .eq("user_id", user_id)
                .limit(1)
                .execute()
                .data) or []
    if not tok:
        abort(400, "No Gmail token; reconnect Gmail first.")

    cd = tok[0]["credentials"]
    creds = Credentials(
        token=cd["token"],
        refresh_token=cd["refresh_token"],
        token_uri=cd["token_uri"],
        client_id=cd["client_id"],
        client_secret=cd["client_secret"],
        scopes=cd["scopes"],
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(GoogleRequest())

    service = build("gmail", "v1", credentials=creds, cache_discovery=False)

    mime = MIMEText(html_body, "html")
    mime["To"]      = ""
    mime["Subject"] = f"Draft Lease: {data['property_name']} → {data['tenant_name']}"
    raw = base64.urlsafe_b64encode(mime.as_bytes()).decode()
    draft = {"message": {"raw": raw}}
    created = service.users().drafts().create(userId="me", body=draft).execute()

    app.logger.info(f"Gmail Draft {created['id']} created for user {user_id}")

    return redirect(f"/dashboard?user_id={user_id}")

@app.route("/admin")
def admin():
    return render_template("admin.html")

@app.route("/api/admin/users")
def api_admin_users():
    users = supabase.table("profiles").select("*").execute().data or []
    today = date.today().isoformat()
    results = []
    for user in users:
        sent = supabase.table("emails") \
            .select("sent_at") \
            .eq("user_id", user["id"]) \
            .eq("status", "sent") \
            .execute().data or []
        count = len([e for e in sent if e["sent_at"] and e["sent_at"].startswith(today)])
        results.append({
            "id": user["id"],
            "name": user["full_name"],
            "email": user["email"],
            "enabled": user.get("ai_enabled", True),
            "emails_today": count
        })
    return jsonify(results)

@app.route("/api/admin/toggle_status", methods=["POST"])
def api_toggle_status():
    user_id = request.json.get("user_id")
    enable = request.json.get("enable", True)
    supabase.table("profiles").update({"ai_enabled": enable}).eq("id", user_id).execute()
    return jsonify({"success": True})

@app.route("/debug_env")
def debug_env():
    return {
        "GOOGLE_CLIENT_ID": os.environ.get("GOOGLE_CLIENT_ID"),
        "REDIRECT_URI": os.environ.get("REDIRECT_URI"),
        "EDGE_BASE_URL": os.environ.get("EDGE_BASE_URL")
    }

from datetime import datetime

# Remove the demo rate limiting decorator and replace with proper plan-based checking
@app.route("/process", methods=["GET"])
def trigger_process():
    token = request.args.get("token")
    if token != os.environ.get("PROCESS_SECRET_TOKEN"):
        return jsonify({"error": "Unauthorized"}), 401

    # ── 0) DAILY RESET CHECK ──
    today_str = date.today().isoformat()
    rl_row = SUPABASE_SERVICE.table("rate_limit_reset") \
        .select("last_reset") \
        .eq("id", "global") \
        .single() \
        .execute().data or {}
    last_date = rl_row.get("last_reset", "")[:10]

    if last_date != today_str:
        app.logger.info("🔄 New day detected – clearing emails table")
        SUPABASE_SERVICE.table("emails") \
            .delete() \
            .neq("id", "00000000-0000-0000-0000-000000000000") \
            .execute()
        SUPABASE_SERVICE.table("rate_limit_reset") \
            .update({"last_reset": datetime.now(timezone.utc).isoformat()}) \
            .eq("id", "global") \
            .execute()

    # ── 1) Fetch ready_to_send emails ──
    ready = (
        supabase.table("emails")
        .select("id, user_id, sender_email, recipient_email, processed_content, subject")
        .eq("status", "ready_to_send")
        .execute()
        .data or []
    )
    
    if not ready:
        # Also check other queues
        gen = supabase.table("emails").select("id, user_id").eq("status", "processing").execute().data or []
        per = supabase.table("emails").select("id, user_id").eq("status", "ready_to_personalize").execute().data or []
        prop = supabase.table("emails").select("id, user_id").eq("status", "awaiting_proposal").execute().data or []
        
        if not (gen or per or prop):
            app.logger.info("⚡ No emails to process — returning 204")
            return "", 204
    
    all_processed, sent, drafted, failed = [], [], [], []
    
    # ── 2) Process AI queues with rate limiting ──
    queues = [
        ("processing", "generate-response"),
        ("ready_to_personalize", "personalize-template"),
        ("awaiting_proposal", "generate-proposal")
    ]
    
    for queue_status, endpoint in queues:
        emails = supabase.table("emails").select("id, user_id").eq("status", queue_status).execute().data or []
        
        if emails:
            # Group by user_id
            emails_by_user = defaultdict(list)
            for email in emails:
                user_id = email.get("user_id")
                if user_id:
                    emails_by_user[user_id].append(email["id"])
            
            for user_id, email_ids in emails_by_user.items():
                # Check rate limit for this batch
                allowed, remaining, message = rate_limiter.check_rate_limit(
                    user_id, 
                    'emails', 
                    len(email_ids)
                )
                
                if not allowed:
                    app.logger.warning(f"Rate limit exceeded for user {user_id} in {queue_status}: {message}")
                    supabase.table("emails").update({
                        "status": "rate_limited",
                        "error_message": f"Plan limit exceeded: {message}"
                    }).in_("id", email_ids).execute()
                    failed.extend(email_ids)
                else:
                    # Call edge function for this batch
                    if call_edge(f"/functions/v1/clever-service/{endpoint}", {"email_ids": email_ids}):
                        if endpoint == "generate-proposal":
                            # These become ready_to_send
                            supabase.table("emails").update({"status": "ready_to_send"}).in_("id", email_ids).execute()
                        elif endpoint == "personalize-template":
                            # These become awaiting_proposal
                            supabase.table("emails").update({"status": "awaiting_proposal"}).in_("id", email_ids).execute()
                        # generate-response automatically updates status
                        all_processed.extend(email_ids)
                    else:
                        supabase.table("emails").update({
                            "status": "error",
                            "error_message": f"{endpoint} failed"
                        }).in_("id", email_ids).execute()
                        failed.extend(email_ids)

    # ── 3) Fetch ready_to_send emails again (including newly generated ones) ──
    ready = (
        supabase.table("emails")
        .select("id, user_id, sender_email, recipient_email, processed_content, subject")
        .eq("status", "ready_to_send")
        .execute()
        .data or []
    )
    
    if not ready:
        return jsonify({
            "processed": all_processed,
            "sent": sent,
            "drafted": drafted,
            "failed": failed,
            "message": "Only AI processing completed, no emails ready to send"
        }), 200

    # Group ready emails by user_id for rate limiting
    ready_by_user = defaultdict(list)
    for rec in ready:
        user_id = rec.get("user_id")
        if user_id:
            ready_by_user[user_id].append(rec)

    # ── 4) SEND via SMTP with rate limiting ──
    for user_id, user_emails in ready_by_user.items():
        # Check rate limit for this user's batch of emails
        allowed, remaining, message = rate_limiter.check_rate_limit(
            user_id, 
            'emails', 
            len(user_emails)
        )
        
        if not allowed:
            app.logger.warning(f"Rate limit exceeded for user {user_id}: {message}")
            # Mark all emails as rate_limited for this user
            for rec in user_emails:
                supabase.table("emails").update({
                    "status": "rate_limited",
                    "error_message": f"Plan limit exceeded: {message}"
                }).eq("id", rec["id"]).execute()
                failed.append(rec["id"])
            continue  # Skip processing for this user

        # Process each email for this user
        for rec in user_emails:
            em_id = rec["id"]
            sender = rec["sender_email"]
            inbox = rec["recipient_email"]
            subject = rec.get("subject", "Your Email")

            # Default credentials placeholders
            smtp_email = None
            smtp_password = None
            smtp_host = "smtp.gmail.com"
            smtp_port = 465
            user_signature = ""
            user_display_name = ""
            generate_leases = False
            
            # --- CHECK IF USING ADMIN CATCH-ALL FORWARDING ---
            # If inbox matches our admin catch-all pattern (e.g. replyzeai.inbound+something@...)
            if "replyzeai.inbound" in inbox.lower():
                # Use Admin Credentials from Environment
                if not ADMIN_INBOUND_EMAIL or not ADMIN_INBOUND_PASSWORD:
                    supabase.table("emails").update({
                        "status": "error",
                        "error_message": "Admin SMTP credentials not configured"
                    }).eq("id", em_id).execute()
                    failed.append(em_id)
                    continue

                # Fetch user profile purely for Signature/Name (using user_id from email record)
                prof = supabase.table("profiles") \
                    .select("display_name, signature, generate_leases") \
                    .eq("id", user_id) \
                    .single().execute().data
                
                if prof:
                    user_display_name = prof.get("display_name", "")
                    user_signature = prof.get("signature", "")
                    generate_leases = prof.get("generate_leases", False)

                # Set credentials to Admin
                smtp_email = inbox # We send FROM the alias address (admin+user@gmail.com)
                smtp_password = ADMIN_INBOUND_PASSWORD
                # Note: smtp_email variable is used as "From" address. 
                # Gmail allows sending as 'admin+alias' if authenticated as 'admin'.

            else:
                # --- STANDARD USER SMTP FLOW ---
                prof = supabase.table("profiles") \
                    .select("smtp_email, smtp_enc_password, smtp_host, smtp_port, display_name, signature, generate_leases") \
                    .eq("smtp_email", inbox) \
                    .execute().data

                if not prof:
                    supabase.table("emails").update({
                        "status": "error",
                        "error_message": f"No SMTP account matches recipient_email {inbox}"
                    }).eq("id", em_id).execute()
                    failed.append(em_id)
                    continue
                
                # Decrypt password
                try:
                    smtp_password = fernet.decrypt(prof["smtp_enc_password"].encode()).decode()
                except Exception as e:
                    supabase.table("emails").update({
                        "status": "error",
                        "error_message": f"SMTP decrypt failed: {str(e)}"
                    }).eq("id", em_id).execute()
                    failed.append(em_id)
                    continue

                smtp_email = prof.get("smtp_email")
                smtp_host = prof.get("smtp_host", "smtp.gmail.com")
                smtp_port = int(prof.get("smtp_port", 465))
                user_signature = prof.get("signature", "")
                user_display_name = prof.get("display_name", "")
                generate_leases = prof.get("generate_leases", False)

            # --- PREPARE EMAIL CONTENT ---
            body_html = (rec.get("processed_content") or "").replace("\n", "<br>")
            if user_signature:
                body_html += f"<br><br>{user_signature}"
            if user_display_name:
                body_html = body_html.replace("[Your Name]", user_display_name)
            final_html = f"<html><body><p>{body_html}</p></body></html>"
            
            # Subject logic
            final_subject = "Lease Agreement Draft" if generate_leases else f"RE: {subject}"
            
            # ─── SEND EMAIL ───
            try:
                send_email_smtp(
                    from_email=smtp_email, # This will be either user's email OR admin+alias
                    from_password=smtp_password,
                    to_email=sender,
                    subject=final_subject,
                    body=final_html,
                    smtp_host=smtp_host,
                    smtp_port=smtp_port
                )
                
                status_to = "drafted" if generate_leases else "sent"
                
                supabase.table("emails").update({
                    "status": status_to,
                    "sent_at": datetime.utcnow().isoformat(),
                    "recipient_email": inbox
                }).eq("id", em_id).execute()
                
                # Increment usage AFTER successful send
                rate_limiter._increment_usage(user_id, 'emails', 1)
                
                if status_to == "sent":
                    sent.append(em_id)
                else:
                    drafted.append(em_id)
                all_processed.append(em_id)
                
            except Exception as e:
                supabase.table("emails").update({
                    "status": "error",
                    "error_message": f"SMTP send failed: {str(e)}"
                }).eq("id", em_id).execute()
                failed.append(em_id)

    return jsonify({
        "processed": all_processed,
        "sent": sent,
        "drafted": drafted,
        "failed": failed
    }), 200



#---------------------------------------------------------------------------------------------------------------------------



#-----------------------------------------------------------------------------------------------------

@app.route("/transaction/<txn_id>/ready", methods=["POST"])
def mark_ready(txn_id):
    supabase.table("transactions").update({"ready_for_kit": True}).eq("id", txn_id).execute()
    return "", 204

@app.route("/autopilot/batch", methods=["POST"])
@check_rate_limit('kits')
def batch_autopilot():
    # Decrement kit count
#    ip = request.remote_addr
 #   demo_rate_limits[ip]['kits'] -= 1
    
    txns = supabase.table("transactions").select("*").eq("ready_for_kit", True).eq("kit_generated", False).execute().data or []
    results = []
    for t in txns:
        payload = {
          "transaction_type": t["transaction_type"],
          "data": {
            "id": t["id"],
            "buyer": t["buyer"],
            "seller": t["seller"],
            "date": t["date"],
            "purchase_price": t["purchase_price"],
            "closing_date": t.get("closing_date"),
            "closing_location": t.get("closing_location")
          }
        }
        resp = requests.post(f"{os.environ.get('BASE_URL')}/autopilot/trigger", json=payload)
        results.append({"id": t["id"], "status": resp.status_code})
        if resp.ok:
            supabase.table("transactions").update({"kit_generated": True}).eq("id", t["id"]).execute()
    return jsonify(results), 200

@app.route("/dashboard/autopilot")
def dashboard_autopilot():
    user_id = request.args.get("user_id") or abort(401)
    txn_id  = request.args.get("txn_id")
    transactions = supabase.table("transactions").select("*").eq("user_id", user_id).execute().data or []
    current_txn = None
    if txn_id:
        resp = supabase.table("transactions").select("*").eq("id", txn_id).execute()
        current_txn = resp.data[0] if resp.data else None
    return render_template("partials/autopilot.html", user_id=user_id, transactions=transactions, current_transaction=current_txn)

@app.route("/transactions/new", methods=["POST"])
def create_transaction():
    import uuid
    import traceback

    user_id = request.args.get("user_id") or request.form.get("user_id")
    if not user_id:
        return jsonify({"status": "error", "message": "Missing user_id"}), 401

    # Check rate limit for leads/transactions
    allowed, remaining, message = rate_limiter.check_rate_limit(user_id, 'leads', 1)
    
    if not allowed:
        return jsonify({
            "status": "error",
            "message": f"Lead creation limit exceeded: {message}"
        }), 429

    new_id = str(uuid.uuid4())

    # 🔐 Validate required fields (lowercase unified names)
    required = ["buyer_name", "seller_name", "property_address", "agreement_date"]
    missing = [f for f in required if not request.form.get(f)]
    if missing:
        app.logger.warning(f"⚠️ Missing required fields: {missing}")
        return jsonify({
            "status": "error",
            "message": f"Missing required fields: {', '.join(missing)}"
        }), 400

    # ✅ All accepted lowercase fields from gamified form
    accepted_fields = [
        "transaction_type", "property_address", "city", "state", "name_of_property",
        "description_of_property", "square_feet", "legal_description",
        "apartment_address", "premises_description",

        "buyer_name", "buyer_address", "seller_name", "seller_address", "agency_name",

        "purchase_price", "deposit_amount", "agreement_date", "broker_name",
        "commission_amount", "brokerage_fee", "broker_payday",

        "closing_date", "occupy_property_date", "mortgage_amount", "mortgage_years",
        "interest_rate", "inspection_days", "possession_date",

        "rent_type", "agreed_rent", "maintenance_terms",

        "landlord_phone", "tenant_phone", "landlord_email", "tenant_email",

        "structure_age", "location", "county", "additional_explanations",

        "buyer_signature", "seller_signature", "time"
    ]

    # Build the payload, turning empty strings into None
    payload = {"id": new_id, "user_id": user_id}
    for field in accepted_fields:
        val = request.form.get(field)
        payload[field] = None if val == "" else val

    try:
        app.logger.info(f"🚀 Inserting transaction with ID {new_id}")
        app.logger.debug(f"Payload: {payload}")
        resp = supabase.table("transactions").insert(payload).execute()
        inserted = resp.data[0]
        
        # Increment usage counter for successful transaction creation
        rate_limiter._increment_usage(user_id, 'leads', 1)
        
    except Exception as e:
        app.logger.error("❌ Transaction insert failed")
        app.logger.error(traceback.format_exc())
        return jsonify({
            "status": "error",
            "message": f"Insertion failed: {str(e)}"
        }), 500

    # ✅ Success response with htmx trigger
    feedback = (
        f'<div class="alert alert-success">🎉 Transaction <strong>{inserted["id"]}</strong> created.</div>'
        + '<script>htmx.trigger(document.querySelector(\'[hx-get*="/dashboard/autopilot"]\'), "click")</script>'
    )
    return feedback, 200

# Add this to your main app file (e.g., app.py)
# Add these imports at the top of your app.py
import re
import dns.resolver

# Add this route to your app.py


def extract_domain(email):
    """Extract domain from email address"""
    pattern = r'@([\w\.-]+)'
    match = re.search(pattern, email)
    if match:
        return match.group(1).lower()
    return None




#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

import csv
from io import TextIOWrapper
from openpyxl import load_workbook

@app.route("/import_leads", methods=["GET", "POST"])
@check_rate_limit('leads')
def import_leads():
    user_id = _require_user()
    
    if request.method == "GET":
        return render_template("import_leads.html", user_id=user_id)
    
    # Handle POST request
    try:
        # Debug logging
        app.logger.info(f"Import leads request received: {request.files}")
        
        if 'file' not in request.files:
            app.logger.error("No file in request")
            return jsonify({"error": "No file uploaded"}), 400
        
        file = request.files['file']
        if file.filename == '':
            app.logger.error("Empty filename")
            return jsonify({"error": "No file selected"}), 400
        
        # Check file extension
        if file.filename.endswith('.csv'):
            # Process CSV file
            csv_file = TextIOWrapper(file, encoding='utf-8')
            reader = csv.DictReader(csv_file)
            rows = list(reader)
            app.logger.info(f"CSV columns: {reader.fieldnames}")
        elif file.filename.endswith(('.xlsx', '.xls')):
            # Process Excel file
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1] if cell.value]
            
            # Get data rows
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                # Remove None values
                rows.append({k: v for k, v in row_data.items() if v is not None})
            
            app.logger.info(f"Excel columns: {headers}")
        else:
            app.logger.error(f"Invalid file type: {file.filename}")
            return jsonify({"error": "Invalid file type. Please upload CSV or Excel."}), 400
        
        # Check if we have any rows
        if not rows:
            app.logger.error("No data rows found in file")
            return jsonify({"error": "No data found in file"}), 400
        
        # Check required columns (more flexible approach)
        required_columns = ['name', 'Last name', 'Recipient', 'city', 'brokerage', 'service', 'Email Sent']
        available_columns = list(rows[0].keys())
        
        app.logger.info(f"Available columns: {available_columns}")
        app.logger.info(f"Required columns: {required_columns}")
        
        missing_columns = [col for col in required_columns if col not in available_columns]
        if missing_columns:
            app.logger.error(f"Missing columns: {missing_columns}")
            return jsonify({
                "error": f"Missing required columns: {', '.join(missing_columns)}. "
                        f"Available columns: {', '.join(available_columns)}"
            }), 400
        
        # Process each row
        success_count = 0
        error_count = 0
        imported_leads = []
        
        for i, row in enumerate(rows):
            try:
                # Parse email_sent date
                email_sent_str = row.get('Email Sent', '')
                email_sent = None
                
                try:
                    if email_sent_str:
                        # Try different date formats
                        if isinstance(email_sent_str, str):
                            try:
                                email_sent = datetime.strptime(email_sent_str, '%Y-%m-%d %H:%M:%S')
                            except ValueError:
                                try:
                                    email_sent = datetime.strptime(email_sent_str, '%Y-%m-%d')
                                except ValueError:
                                    # Try to parse Excel serial date numbers
                                    try:
                                        if isinstance(email_sent_str, (int, float)):
                                            email_sent = datetime(1899, 12, 30) + timedelta(days=email_sent_str)
                                        else:
                                            email_sent = datetime.utcnow()
                                    except:
                                        email_sent = datetime.utcnow()
                        else:
                            # Assume it's already a datetime object
                            email_sent = email_sent_str
                    else:
                        email_sent = datetime.utcnow()
                except Exception as e:
                    app.logger.warning(f"Error parsing date {email_sent_str}: {e}")
                    email_sent = datetime.utcnow()
                
                # Prepare lead data
                lead_data = {
                    'user_id': user_id,
                    'first_name': str(row.get('name', '')).strip(),
                    'last_name': str(row.get('Last name', '')).strip(),
                    'email': str(row.get('Recipient', '')).strip().lower(),
                    'city': str(row.get('city', '')).strip(),
                    'brokerage': str(row.get('brokerage', '')).strip(),
                    'service': str(row.get('service', '')).strip(),
                    'status': 'new',
                    'email_sent': email_sent.isoformat() if hasattr(email_sent, 'isoformat') else email_sent,
                    'created_at': datetime.utcnow().isoformat()
                }
                
                # Validate required fields
                if not lead_data['email'] or '@' not in lead_data['email']:
                    app.logger.warning(f"Row {i+1}: Invalid email address '{lead_data['email']}'")
                    error_count += 1
                    continue
                
                if not lead_data['first_name'] and not lead_data['last_name']:
                    app.logger.warning(f"Row {i+1}: Missing both first and last name")
                    error_count += 1
                    continue
                
                # Insert lead
                response = supabase.table('leads').insert(lead_data).execute()
                
                if response.data:
                    success_count += 1
                    lead_id = response.data[0]['id']
                    imported_leads.append(lead_id)
                    
                    # Schedule follow-ups
                    try:
                        # Send immediate follow-up (step 0)
                        follow_up_content = generate_follow_up_content(lead_id, 0)
                        
                        if follow_up_content:
                            # Get lead details
                            lead = supabase.table('leads').select('*').eq('id', lead_id).single().execute().data
                            
                            # Send email using Gmail API
                            success, message = send_email_gmail(
                                user_id,
                                lead['email'],
                                "Follow-up from your inquiry",
                                follow_up_content
                            )
                            
                            # FIXED INDENTATION: This was the main issue
                            if success:
                                # Create follow-up record
                                follow_up_data = {
                                    'lead_id': lead_id,
                                    'sequence_step': 0,
                                    'generated_content': follow_up_content,
                                    'status': 'sent',
                                    'sent_at': datetime.utcnow().isoformat()
                                }
                                supabase.table('lead_follow_ups').insert(follow_up_data).execute()
                            else:
                                app.logger.error(f"Failed to send immediate follow-up for lead {lead_id}: {message}")
                    
                    except Exception as e:
                        app.logger.error(f"Error sending immediate follow-up for lead {lead_id}: {str(e)}")
                    
                    # Schedule the rest of the follow-up sequence (unchanged)
                    for step, seq in enumerate(FOLLOW_UP_SEQUENCE[1:], start=1):
                        scheduled_at = email_sent + timedelta(days=seq['delay_days'])
                        follow_up_data = {
                            'lead_id': lead_id,
                            'sequence_step': step,
                            'scheduled_at': scheduled_at.isoformat(),
                            'status': 'pending'
                        }
                        supabase.table('lead_follow_ups').insert(follow_up_data).execute()
                
                else:
                    error_count += 1
                    app.logger.error(f"Failed to insert lead: {response}")
            
            except Exception as e:
                error_count += 1
                app.logger.error(f"Error processing row {i+1}: {e}", exc_info=True)
        
        # Log summary
        app.logger.info(f"Import completed: {success_count} succeeded, {error_count} failed")
        
        return jsonify({
            "message": f"Leads imported successfully. {success_count} succeeded, {error_count} failed.",
            "imported_count": success_count,
            "failed_count": error_count
        }), 200
    
    except Exception as e:
        app.logger.error(f"Error importing leads: {str(e)}", exc_info=True)
        return jsonify({"error": f"Failed to import leads: {str(e)}"}), 500
#------------------------------------------------------------------------------------------------------------------
def generate_follow_up_content(lead_id, sequence_step):
    """Generate follow-up content using AI with context of previous communications"""
    try:
        app.logger.info(f"Starting follow-up generation for lead {lead_id}, step {sequence_step}")
        
        # Get lead details
        lead_resp = supabase.table("leads").select("*").eq("id", lead_id).single().execute()
        if not lead_resp.data:
            app.logger.error(f"Lead {lead_id} not found")
            return None
            
        lead = lead_resp.data
        app.logger.info(f"Found lead: {lead['email']}")
        
        # Get previous emails from emails table
        previous_emails = supabase.table("emails") \
            .select("subject, original_content, processed_content, sent_at") \
            .eq("sender_email", lead["email"]) \
            .order("sent_at", desc=True) \
            .limit(5) \
            .execute().data or []
        
        # Get previous follow-ups from lead_follow_ups table
        previous_follow_ups = supabase.table("lead_follow_ups") \
            .select("generated_content, sent_at, sequence_step") \
            .eq("lead_id", lead_id) \
            .eq("status", "sent") \
            .lt("sequence_step", sequence_step) \
            .order("sent_at", desc=True) \
            .execute().data or []
        
        app.logger.info(f"Found {len(previous_emails)} previous emails and {len(previous_follow_ups)} previous follow-ups")
        
        # Build context for AI
        context = f"""
        Lead: {lead['first_name']} {lead['last_name']}
        Company: {lead['brokerage']}
        Service: {lead['service']}
        Location: {lead['city']}
        
        Previous communications:
        """
        
        # Add emails from emails table
        for i, email in enumerate(previous_emails):
            context += f"\nEmail {i+1} ({email.get('sent_at', '')}):\n"
            context += f"Subject: {email.get('subject', 'No subject')}\n"
            content = email.get('original_content') or email.get('processed_content', '')
            context += f"Content: {content[:200]}...\n" if len(content) > 200 else f"Content: {content}\n"
        
        # Add follow-ups from lead_follow_ups table
        for i, follow_up in enumerate(previous_follow_ups, start=len(previous_emails)+1):
            context += f"\nFollow-up {i} (Day {FOLLOW_UP_SEQUENCE[follow_up['sequence_step']]['delay_days']}, {follow_up.get('sent_at', '')}):\n"
            content = follow_up.get('generated_content', '')
            context += f"Content: {content[:200]}...\n" if len(content) > 200 else f"Content: {content}\n"
        
        if not previous_emails and not previous_follow_ups:
            context += "\nNo previous communications found. This is the first contact.\n"
        
        context += f"\n\nWrite a friendly, professional follow-up email for day {FOLLOW_UP_SEQUENCE[sequence_step]['delay_days']}."
        context += " Reference previous communications if relevant. Keep it concise and focused on providing value."
        
        app.logger.info(f"Built context for AI: {context[:500]}...")
        
        # Call your AI API
        payload = {
            "context": context,
            "type": "follow_up",
            "sequence_step": sequence_step,
            "lead_id": lead_id
        }
        
        app.logger.info(f"Calling edge function with payload: {payload}")
        
        # Use your existing Edge Function call pattern
        # Modify call_edge to return the response content instead of just success/failure
        response = call_edge("/functions/v1/generate-follow-up", payload, return_response=True)
        
        if response and response.status_code == 200:
            content = response.json().get("content")
            app.logger.info(f"Successfully generated follow-up for lead {lead_id}")
            return content
        else:
            app.logger.error(f"Failed to generate follow-up content for lead {lead_id}")
            return None
            
    except Exception as e:
        app.logger.error(f"Error generating follow-up content: {str(e)}", exc_info=True)
        return None
#-------------------------------------------------------------------------------------------------------------------------------------------------


# Update the process_follow_ups route to use Gmail API
@app.route("/process_follow_ups", methods=["GET"])
def process_follow_ups():
    # Check for secret token
    token = request.args.get("token")
    if token != os.environ.get("PROCESS_SECRET_TOKEN"):
        return jsonify({"error": "Unauthorized"}), 401
    
    try:
        # Get due follow-ups
        now = datetime.now(timezone.utc).isoformat()
        due_follow_ups = supabase.table("lead_follow_ups") \
            .select("*, leads(*)") \
            .lte("scheduled_at", now) \
            .eq("status", "pending") \
            .execute().data
        
        if not due_follow_ups:
            return jsonify({"message": "No follow-ups to process"}), 200
        
        results = {"processed": [], "failed": []}
        
        # Group follow-ups by user_id for rate limiting
        follow_ups_by_user = defaultdict(list)
        for follow_up in due_follow_ups:
            user_id = follow_up["leads"]["user_id"]
            follow_ups_by_user[user_id].append(follow_up)
        
        # Process each user's follow-ups
        for user_id, user_follow_ups in follow_ups_by_user.items():
            # Check rate limit for this user
            allowed, remaining, message = rate_limiter.check_rate_limit(
                user_id, 
                'cold_emails',  # Using 'emails' resource for follow-ups
                len(user_follow_ups)
            )
            
            if not allowed:
                app.logger.warning(f"Rate limit exceeded for user {user_id}: {message}")
                # Mark all follow-ups as rate_limited for this user
                for follow_up in user_follow_ups:
                    supabase.table("lead_follow_ups") \
                        .update({
                            "status": "failed", 
                            "error_message": f"Plan limit exceeded: {message}"
                        }) \
                        .eq("id", follow_up["id"]) \
                        .execute()
                    results["failed"].append(follow_up["id"])
                continue  # Skip processing for this user
            
            # Process each follow-up for this user
            for follow_up in user_follow_ups:
                try:
                    # Generate content using AI
                    content = generate_follow_up_content(follow_up["lead_id"], follow_up["sequence_step"])
                    if content:
                        lead_email = follow_up["leads"]["email"]
                        
                        # Get user's SMTP credentials
                        user_profile = supabase.table("profiles") \
                            .select("smtp_email, smtp_enc_password, smtp_host, smtp_port, display_name") \
                            .eq("id", user_id) \
                            .single().execute().data
                        
                        if not user_profile or not user_profile.get("smtp_email"):
                            supabase.table("lead_follow_ups") \
                                .update({
                                    "status": "failed", 
                                    "error_message": "No SMTP account configured for user"
                                }) \
                                .eq("id", follow_up["id"]) \
                                .execute()
                            results["failed"].append(follow_up["id"])
                            continue
                        
                        # Decrypt SMTP password
                        try:
                            smtp_password = fernet.decrypt(user_profile["smtp_enc_password"].encode()).decode()
                        except Exception as e:
                            supabase.table("lead_follow_ups") \
                                .update({
                                    "status": "failed", 
                                    "error_message": f"SMTP password decryption failed: {str(e)}"
                                }) \
                                .eq("id", follow_up["id"]) \
                                .execute()
                            results["failed"].append(follow_up["id"])
                            continue
                        
                        # Prepare email content
                        subject = f"Follow-up: {follow_up['leads'].get('first_name', '')} {follow_up['leads'].get('last_name', '')}"
                        
                        # Build HTML body
                        body_html = content.replace("\n", "<br>")
                        if user_profile.get("display_name"):
                            # Replace [Your Name] placeholder if exists
                            body_html = body_html.replace("[Your Name]", user_profile["display_name"])
                        
                        final_html = f"<html><body><p>{body_html}</p></body></html>"
                        
                        # Send using SMTP
                        try:
                            rate_limiter._increment_usage(user_id, 'cold_emails', 1)
                            # DEBUG: Check if increment worked
                            app.logger.info(f"Increment usage called for user {user_id}")
                            send_email_smtp(
                                from_email=user_profile["smtp_email"],
                                from_password=smtp_password,
                                to_email=lead_email,
                                subject=subject,
                                body=final_html,
                                smtp_host=user_profile.get("smtp_host", "smtp.gmail.com"),
                                smtp_port=int(user_profile.get("smtp_port", 465))
                            )
                            
                            # Update status
                            supabase.table("lead_follow_ups") \
                                .update({
                                    "status": "sent", 
                                    "generated_content": content,
                                    "sent_at": now
                                }) \
                                .eq("id", follow_up["id"]) \
                                .execute()
                            
                            # Increment usage counter
                            
                            
                            results["processed"].append(follow_up["id"])
                            
                        except Exception as e:
                            supabase.table("lead_follow_ups") \
                                .update({
                                    "status": "failed", 
                                    "error_message": f"SMTP send failed: {str(e)}"
                                }) \
                                .eq("id", follow_up["id"]) \
                                .execute()
                            results["failed"].append(follow_up["id"])
                            
                    else:
                        supabase.table("lead_follow_ups") \
                            .update({"status": "failed", "error_message": "Failed to generate content"}) \
                            .eq("id", follow_up["id"]) \
                            .execute()
                        results["failed"].append(follow_up["id"])
                        
                except Exception as e:
                    app.logger.error(f"Error processing follow-up {follow_up['id']}: {str(e)}")
                    supabase.table("lead_follow_ups") \
                        .update({
                            "status": "failed", 
                            "error_message": f"Processing error: {str(e)}"
                        }) \
                        .eq("id", follow_up["id"]) \
                        .execute()
                    results["failed"].append(follow_up["id"])
        
        return jsonify(results), 200
        
    except Exception as e:
        app.logger.error(f"Error in process_follow_ups: {str(e)}")
        return jsonify({"error": str(e)}), 500

#----------------------------------------------------------------------------------------------------------------------------------------
# Add to app.py

@app.route("/api/generate-complete-kit", methods=["POST"])
def generate_complete_kit():
    """Generate a complete closing kit with all document types"""
    data = request.get_json()
    ip = request.remote_addr
    
    # Check rate limits
    if (ip not in demo_rate_limits or 
        'kits' not in demo_rate_limits[ip] or 
        demo_rate_limits[ip]['kits'] <= 0):
        
        return jsonify({"error": "Closing kit limit exceeded"}), 429
    
    try:
        # Decrement the limit
        demo_rate_limits[ip]['kits'] -= 1
        
        # Generate all document types
        docs = []
        templates = [
            ("loi_template.docx", "LOI"),
            ("psa_template.docx", "PSA"),
            ("purchase_offer_template.docx", "PURCHASE_OFFER"),
            ("agency_disclosure_template.docx", "AGENCY_DISCLOSURE"),
            ("real_estate_purchase_template.docx", "REAL_ESTATE_PURCHASE"),
            ("lease_template.docx", "LEASE"),
            ("seller_disclosure_template.docx", "SELLER_DISCLOSURE"),
        ]
        
        # Create temporary directory for documents
        import tempfile
        import uuid
        tmpdir = tempfile.mkdtemp()
        
        for template_name, prefix in templates:
            try:
                tpl = DocxTemplate(f"templates/transaction_autopilot/{template_name}")
                
                # Map form data to template variables
                template_data = map_form_data_to_template(data, prefix.lower())
                tpl.render(template_data)
                
                out_name = f"{prefix}_{data.get('id', 'demo')}_{uuid.uuid4().hex[:6]}.docx"
                out_path = os.path.join(tmpdir, out_name)
                tpl.save(out_path)
                docs.append(out_path)
            except Exception as e:
                app.logger.error(f"Error generating {template_name}: {str(e)}")
                continue
        
        # Bundle into ZIP
        zip_io = BytesIO()
        with zipfile.ZipFile(zip_io, "w") as zf:
            for doc_path in docs:
                zf.write(doc_path, arcname=os.path.basename(doc_path))
        
        zip_io.seek(0)
        
        # Clean up temporary files
        for doc_path in docs:
            try:
                os.remove(doc_path)
            except:
                pass
                
        try:
            os.rmdir(tmpdir)
        except:
            pass
        
        # Return the ZIP file
        return send_file(
            zip_io,
            as_attachment=True,
            download_name=f"complete_closing_kit_{data.get('id', 'demo')}.zip",
            mimetype="application/zip"
        )
        
    except Exception as e:
        app.logger.error(f"Error generating closing kit: {str(e)}")
        return jsonify({"error": str(e)}), 500

def map_form_data_to_template(form_data, doc_type):
    """Map form data to appropriate template variables based on document type"""
    mapped_data = form_data.copy()
    
    # Add common mappings
    mapped_data['transaction_id'] = form_data.get('id', '')
    mapped_data['current_date'] = datetime.now().strftime('%B %d, %Y')
    
    # Document-specific mappings
    if doc_type == 'loi':
        mapped_data['letter_date'] = datetime.now().strftime('%B %d, %Y')
        mapped_data['buyer_signature'] = form_data.get('buyer_signature', '')
        mapped_data['seller_signature'] = form_data.get('seller_signature', '')
    
    elif doc_type == 'psa':
        mapped_data['effective_date'] = form_data.get('agreement_date', '')
        mapped_data['closing_date'] = form_data.get('closing_date', '')
        mapped_data['purchase_price'] = f"${float(form_data.get('purchase_price', 0)):,.2f}"
    
    elif doc_type == 'lease':
        mapped_data['lease_term'] = form_data.get('rent_type', '')
        mapped_data['monthly_rent'] = f"${float(form_data.get('agreed_rent', 0)):,.2f}"
        mapped_data['security_deposit'] = f"${float(form_data.get('deposit_amount', 0)):,.2f}"
    
    return mapped_data

# Add a route to check rate limit status
@app.route("/rate_limit_status")
def rate_limit_status():
    ip = request.remote_addr
    now = datetime.now()
    
    # Check and reset limits if needed (same logic as decorator)
    for resource in ['emails', 'kits', 'leads']:
        if resource == 'emails':
            if (now - demo_rate_limits[ip][resource]['last_reset']).days >= 1:
                demo_rate_limits[ip][resource]['remaining'] = 20
                demo_rate_limits[ip][resource]['last_reset'] = now
        else:
            if (now - demo_rate_limits[ip][resource]['last_reset']).days >= 30:
                if resource == 'kits':
                    demo_rate_limits[ip][resource]['remaining'] = 20
                else:
                    demo_rate_limits[ip][resource]['remaining'] = 25
                demo_rate_limits[ip][resource]['last_reset'] = now
    
    return jsonify({
        'emails_remaining': demo_rate_limits[ip]['emails']['remaining'],
        'kits_remaining': demo_rate_limits[ip]['kits']['remaining'],
        'leads_remaining': demo_rate_limits[ip]['leads']['remaining'],
        'emails_reset': (demo_rate_limits[ip]['emails']['last_reset'] + timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S'),
        'kits_reset': (demo_rate_limits[ip]['kits']['last_reset'] + timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S'),
        'leads_reset': (demo_rate_limits[ip]['leads']['last_reset'] + timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
    })
#-----------------------------------------------------------------------------------------------------------------------------------------
# Add a new route to test Gmail connection
@app.route("/test_gmail_connection", methods=["POST"])
def test_gmail_connection():
    user_id = _require_user()
    
    try:
        service = get_gmail_service(user_id)
        if not service:
            return jsonify({"success": False, "message": "Gmail service not available"})
        
        # Test by getting user profile
        profile = service.users().getProfile(userId='me').execute()
        email_address = profile.get('emailAddress')
        
        return jsonify({
            "success": True, 
            "message": f"Gmail connection successful for {email_address}"
        })
        
    except Exception as e:
        app.logger.error(f"Gmail connection test failed for user {user_id}: {str(e)}")
        return jsonify({"success": False, "message": str(e)})

#---------------------------------------------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------
#show forwarding status:
@app.route("/dashboard/email_forwarding")
def email_forwarding_settings():
    user_id = _require_user()
    
    # Get user profile with forwarding status
    profile = supabase.table("profiles") \
        .select("email, display_name, forwarding_verified, forwarding_verified_at") \
        .eq("id", user_id) \
        .single() \
        .execute().data or {}
    
    # Dynamically generate the user's specific catch-all address
    # e.g., replyzeai.inbound+JohnDoe@gmail.com
    username = normalize_display_name(profile.get("display_name", ""))
    
    # Split the base admin email to insert the alias part
    # Assuming ADMIN_INBOUND_EMAIL is like "name@gmail.com"
    if "@" in ADMIN_INBOUND_EMAIL:
        base, domain = ADMIN_INBOUND_EMAIL.split("@", 1)
        # Create the plus-alias address
        user_specific_forwarding_address = f"{base}+{username}@{domain}"
    else:
        # Fallback if config is weird
        user_specific_forwarding_address = "Contact Support"

    return render_template(
        "partials/email_forwarding.html",
        profile=profile,
        user_id=user_id,
        polling_email=user_specific_forwarding_address # Send the user-specific alias to the template
    )


#------------------------------------------------------------------------------------------------------

@app.route("/check_forwarding_status")
def check_forwarding_status():
    """Check if user has email forwarding enabled"""
    user_id = _require_user()
    
    try:
        # Get user profile with forwarding status
        profile = supabase.table("profiles") \
            .select("forwarding_verified, forwarding_verified_at, email") \
            .eq("id", user_id) \
            .single() \
            .execute().data or {}
        
        if profile.get("forwarding_verified"):
            return jsonify({
                "status": "connected",
                "message": f"Forwarding verified on {profile.get('forwarding_verified_at', '')[:10]}",
                "email": profile.get("email", "")
            })
        else:
            return jsonify({
                "status": "not_connected", 
                "message": "Email forwarding not set up",
                "email": profile.get("email", "")
            })
            
    except Exception as e:
        app.logger.error(f"Error checking forwarding status: {str(e)}")
        return jsonify({
            "status": "error",
            "message": "Error checking forwarding status"
        }), 500

#-------------------------------------------------
#---------the manual dash--------------------
@app.route("/dashboard/manual_email", methods=["GET", "POST"])
def manual_email():
    """Manual email input for users who don't want email forwarding"""
    user_id = _require_user()
    
    if request.method == "POST":
        # Process the manual email
        email_content = request.form.get("email_content", "").strip()
        sender_email = request.form.get("sender_email", "").strip()
        subject = request.form.get("subject", "Inquiry") or "Inquiry"
        
        if not email_content:
            return jsonify({"error": "Email content is required"}), 400
        
        try:
            # Get user's email to use as recipient_email
            user_profile = supabase.table("profiles") \
                .select("email") \
                .eq("id", user_id) \
                .single() \
                .execute().data
            
            user_email = user_profile.get("email") if user_profile else "user@example.com"
            
            # Create email record with ALL required fields
            email_data = {
                "user_id": user_id,
                "sender_email": sender_email or "manual@input.com",
                "recipient_email": user_email,  # REQUIRED FIELD - use user's email
                "original_content": email_content,
                "subject": subject,
                "status": "processing",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "original_user_id": user_id,  # Also set this for consistency
                "is_forwarded": False  # Mark as manual email
            }
            
            # Use SUPABASE_SERVICE (service role) to bypass RLS
            result = SUPABASE_SERVICE.table("emails").insert(email_data).execute()
            
            if result.data:
                # Trigger AI processing
                success = call_edge("/functions/v1/clever-service/generate-response", 
                                  {"email_ids": [result.data[0]["id"]]})
                
                return jsonify({
                    "success": True,
                    "message": "Email submitted for AI processing",
                    "email_id": result.data[0]["id"]
                })
            else:
                return jsonify({"error": "Failed to save email"}), 500
                
        except Exception as e:
            app.logger.error(f"Error processing manual email: {str(e)}")
            return jsonify({"error": f"Failed to process email: {str(e)}"}), 500
    
    # GET request - render the manual email form
    return render_template("partials/manual_email.html", user_id=user_id)

@app.route("/check_manual_email_status/<email_id>")
def check_manual_email_status(email_id):
    """Check status of manually submitted email"""
    user_id = _require_user()
    
    try:
        # Use service role to bypass RLS for reading
        email = SUPABASE_SERVICE.table("emails") \
            .select("id, status, processed_content, error_message, user_id") \
            .eq("id", email_id) \
            .single() \
            .execute().data
        
        if not email:
            return jsonify({"error": "Email not found"}), 404
            
        # Verify the email belongs to the current user
        if email["user_id"] != user_id:
            return jsonify({"error": "Access denied"}), 403
            
        return jsonify({
            "status": email["status"],
            "processed_content": email.get("processed_content"),
            "error_message": email.get("error_message")
        })
        
    except Exception as e:
        app.logger.error(f"Error checking email status: {str(e)}")
        return jsonify({"error": "Failed to check status"}), 500

#----------------------------------------------------------------
#---------------- manual follow ups -------------------------------



def generate_fallback_follow_up(lead, sequence_step):
    """Generate a simple fallback follow-up when AI fails"""
    lead_name = lead['first_name'] or "there"
    days = FOLLOW_UP_SEQUENCE[sequence_step]['delay_days']
    
    follow_ups = [
        f"Hi {lead_name}, I wanted to follow up on my previous email about commercial real estate opportunities in your area. Are you still interested in exploring options?",
        f"Hello {lead_name}, checking in to see if you've had a chance to consider commercial properties recently. I'm here to help if you have any questions.",
        f"Hi {lead_name}, I'm following up on our previous conversation about commercial real estate. The market has been active lately - would you like me to update you on current opportunities?",
        f"Hello {lead_name}, I wanted to reconnect regarding commercial property options. Have your requirements changed since we last connected?",
        f"Hi {lead_name}, just checking in to see if you're still in the market for commercial space. I've come across some new listings that might interest you.",
        f"Hello {lead_name}, I'm following up on our previous discussion. Is this still a good time to explore commercial real estate opportunities?"
    ]
    
    # Use sequence step to pick appropriate fallback, or random if beyond list
    return follow_ups[sequence_step % len(follow_ups)]

@app.route("/generate_manual_followups", methods=["POST"])
def generate_manual_followups():
    """Generate AI-powered follow-ups for a manual email"""
    user_id = _require_user()
    
    try:
        data = request.get_json()
        sender_email = data.get("sender_email")
        sender_name = data.get("sender_name")
        subject = data.get("subject")
        email_content = data.get("email_content")
        
        if not all([sender_email, sender_name, email_content]):
            return jsonify({"error": "Missing required fields"}), 400
        
        # Create a temporary lead record for follow-up generation
        lead_data = {
            "user_id": user_id,
            "first_name": sender_name.split()[0] if sender_name else "Lead",
            "last_name": " ".join(sender_name.split()[1:]) if sender_name and " " in sender_name else "Contact",
            "email": sender_email,
            "brokerage": "Unknown",  # Default values
            "service": "Commercial Real Estate",
            "city": "Unknown",
            "status": "new",
            "email_sent": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Insert the lead using service role to bypass RLS
        lead_result = SUPABASE_SERVICE.table("leads").insert(lead_data).execute()
        
        if not lead_result.data:
            return jsonify({"error": "Failed to create lead record"}), 500
            
        lead_id = lead_result.data[0]["id"]
        
        # Also store the original email for context using service role
        email_record = {
            "user_id": user_id,
            "sender_email": sender_email,
            "recipient_email": "manual@input.com",  # Placeholder
            "original_content": email_content,
            "subject": subject,
            "status": "manual_input",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "original_user_id": user_id,
            "is_forwarded": False
        }
        
        SUPABASE_SERVICE.table("emails").insert(email_record).execute()
        
        # Generate follow-ups using the same AI system
        follow_ups = []
        for step, seq in enumerate(FOLLOW_UP_SEQUENCE):
            # Generate content for this follow-up step
            content = generate_follow_up_content(lead_id, step)
            
            if content:
                scheduled_at = datetime.now(timezone.utc) + timedelta(days=seq['delay_days'])
                
                follow_up_data = {
                    "lead_id": lead_id,
                    "sequence_step": step,
                    "scheduled_at": scheduled_at.isoformat(),
                    "status": "pending",
                    "generated_content": content
                }
                
                # Store the follow-up using service role
                follow_up_result = SUPABASE_SERVICE.table("lead_follow_ups").insert(follow_up_data).execute()
                
                if follow_up_result.data:
                    follow_ups.append({
                        "id": follow_up_result.data[0]["id"],
                        "lead_id": lead_id,
                        "day": seq['name'],
                        "date": scheduled_at.strftime("%Y-%m-%d"),
                        "content": content,
                        "sequence_step": step,
                        "delay_days": seq['delay_days']
                    })
        
        return jsonify({
            "success": True,
            "lead_id": lead_id,
            "follow_ups": follow_ups,
            "message": f"Generated {len(follow_ups)} AI-powered follow-ups"
        })
        
    except Exception as e:
        app.logger.error(f"Error generating manual follow-ups: {str(e)}", exc_info=True)
        return jsonify({"error": f"Failed to generate follow-ups: {str(e)}"}), 500




#-------------------------------------------------------------
#--------------------manual /auto--------------------------------------
@app.route("/set_email_mode", methods=["POST"])
def set_email_mode():
    """Set user's preferred email mode (auto or manual)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON data provided"}), 400
            
        user_id = data.get("user_id")
        mode = data.get("mode")
        
        if not user_id:
            return jsonify({"error": "Missing user_id"}), 401
            
        if mode not in ["auto", "manual"]:
            return jsonify({"error": "Invalid mode"}), 400
        
        # Update user's email mode preference
        supabase.table("profiles").update({
            "email_mode": mode
        }).eq("id", user_id).execute()
        
        return jsonify({"success": True, "mode": mode})
    except Exception as e:
        app.logger.error(f"Error setting email mode: {str(e)}")
        return jsonify({"error": "Failed to set email mode"}), 500

@app.route("/get_email_mode")
def get_email_mode():
    """Get user's current email mode"""
    try:
        user_id = request.args.get("user_id")
        if not user_id:
            return jsonify({"error": "Missing user_id"}), 401
            
        profile = supabase.table("profiles") \
            .select("email_mode") \
            .eq("id", user_id) \
            .single() \
            .execute().data
        
        return jsonify({"mode": profile.get("email_mode", "auto")})
    except Exception as e:
        app.logger.error(f"Error getting email mode: {str(e)}")
        return jsonify({"mode": "auto"})  # Default to auto


# ── Final entry point ──
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
